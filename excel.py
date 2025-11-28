"""
Модуль для резервного копирования данных в Excel с использованием openpyxl.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import logging

# Настройка логирования
logger = logging.getLogger(__name__)

# Названия файлов и колонок
USERS_BACKUP_FILE = 'users_backup.xlsx'
TRANSACTIONS_BACKUP_FILE = 'transactions_backup.xlsx'

USER_COLUMNS = ['user_id', 'name', 'currency', 'timestamp']
TRANSACTION_COLUMNS = ['user_id', 'type', 'amount', 'category', 'description', 'date']

def _setup_workbook(file_path: str, columns: list):
    """Создает новый Excel файл с заголовками."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(columns)
    # Авто-подбор ширины колонок
    for i, col_name in enumerate(columns, 1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = len(col_name) + 5
    workbook.save(file_path)
    logger.info(f"Создан файл бэкапа: {file_path}")

def backup_user(user_id: int, name: str, currency: str):
    """
    Сохраняет или обновляет информацию о пользователе в Excel-файле.
    """
    file_path = USERS_BACKUP_FILE
    
    try:
        if not os.path.exists(file_path):
            _setup_workbook(file_path, USER_COLUMNS)

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Ищем пользователя
        user_row = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == user_id:
                user_row = row
                break

        new_data = [user_id, name, currency, datetime.now()]

        if user_row:
            # Обновляем существующую строку
            for i, value in enumerate(new_data, 1):
                sheet.cell(row=user_row, column=i, value=value)
            logger.info(f"Данные пользователя {user_id} обновлены в {file_path}")
        else:
            # Добавляем новую строку
            sheet.append(new_data)
            logger.info(f"Новый пользователь {user_id} добавлен в {file_path}")

        workbook.save(file_path)

    except Exception as e:
        logger.error(f"Ошибка при сохранении пользователя {user_id} в Excel: {e}")

def backup_transaction(user_id: int, transaction_type: str, amount: float, category: str, description: str):
    """
    Добавляет запись о новой транзакции в Excel-файл.
    """
    file_path = TRANSACTIONS_BACKUP_FILE
    
    try:
        if not os.path.exists(file_path):
            _setup_workbook(file_path, TRANSACTION_COLUMNS)

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        new_data = [user_id, transaction_type, amount, category, description, datetime.now()]
        sheet.append(new_data)
        
        workbook.save(file_path)
        logger.info(f"Транзакция для пользователя {user_id} сохранена в {file_path}")

    except Exception as e:
        logger.error(f"Ошибка при сохранении транзакции для {user_id} в Excel: {e}")